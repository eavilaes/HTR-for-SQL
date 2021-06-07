import os
from openpyxl import Workbook

import argparse
import json

import cv2
import editdistance
from path import Path

from DataLoaderIAM import DataLoaderIAM, Batch
from Model import Model, DecoderType
from SamplePreprocessor import preprocess

from tqdm import tqdm

NUM_IMGS_PER_WORD = 5
NUM_LINES = 445

class FilePaths:
    "filenames and paths to data"
    fnCharList = '../model/charList.txt'
    fnSummary = '../model/summary.json'
    fnInfer = '../../../HTG/convolutional-handwriting-gan/output/'    # Output folder from the HTG
    fnCorpus = '../data/corpus.txt'



def infer(model, fnImg):
    "recognize text in image provided by file path"
    img = preprocess(cv2.imread(fnImg, cv2.IMREAD_GRAYSCALE), Model.imgSize)
    batch = Batch(None, [img])
    (recognized, probability) = model.inferBatch(batch, True)
    # print(f'Recognized: "{recognized[0]}"')
    # print(f'Probability: {probability[0]}')

    return (recognized, probability)


def main():
    "main function"
    parser = argparse.ArgumentParser()
    parser.add_argument('--decoder', choices=['bestpath', 'beamsearch', 'wordbeamsearch'], default='bestpath',
                        help='CTC decoder')
    parser.add_argument('--dump', help='dump output of NN to CSV file(s)', action='store_true')
    args = parser.parse_args()

    # set chosen CTC decoder
    if args.decoder == 'bestpath':
        decoderType = DecoderType.BestPath
    elif args.decoder == 'beamsearch':
        decoderType = DecoderType.BeamSearch
    elif args.decoder == 'wordbeamsearch':
        decoderType = DecoderType.WordBeamSearch

    # infer text on test image
    model = Model(open(FilePaths.fnCharList).read(), decoderType, mustRestore=True, dump=args.dump)
    
    wb = Workbook() # Work with .xlsx files (Excel)
    data = [[0 for i in range(2)] for j in range(NUM_LINES)] # This will store all the data related to the recognition and probabilities (excel sheets)
    
    err = 0
    # Loop on all images calling infer() to get the output
    for root, dirs, files in os.walk(FilePaths.fnInfer):
        for directory in tqdm(dirs):
            wr = wb.create_sheet(directory)
            wp = wb.create_sheet(directory + "_prob")
            for _, _, files in os.walk(FilePaths.fnInfer + directory):
                cellsR = [[0 for i in range(NUM_IMGS_PER_WORD)] for j in range(len(files)//NUM_IMGS_PER_WORD)]
                cellsP = [[0 for i in range(NUM_IMGS_PER_WORD)] for j in range(len(files)//NUM_IMGS_PER_WORD)]
                for file in files:
                    word = file.split('_')[0][4:]
                    img  = file.split('_')[1][3:].split('.')[0]
                    try:
                        recprob = infer(model, FilePaths.fnInfer + directory + "/" + file) # Infer the image and get the recognized and probability as a tuple
                    except:
                        err += 1
                        # print("Error in line: ", directory, file, "Total errors: ", err)
                        pass
                    cellsR [int(word)][int(img)] = recprob[0][0]
                    cellsP [int(word)][int(img)] = recprob[1][0]
                for w in range(len(cellsR)):
                    wr.append(cellsR[w])
                    wp.append(cellsP[w])
                wb.save("inferOutput.xlsx")

                data[int(directory[4:])-1][0] = cellsR
                data[int(directory[4:])-1][1] = cellsP

        break
        # path = root.split(os.sep)
        # #print('---', os.path.basename(root)) # Folder (line) name = os.path.basename(root)
        # wr = wb.create_sheet(os.path.basename(root)) # Create new sheet for each line (sentence)
        # wp = wb.create_sheet(os.path.basename(root) + "_prob") # Create new sheet for probabilities

        # cellsR = [[0 for i in range(NUM_IMGS_PER_WORD)] for j in range(len(files)//NUM_IMGS_PER_WORD)]
        # cellsP = [[0 for i in range(NUM_IMGS_PER_WORD)] for j in range(len(files)//NUM_IMGS_PER_WORD)]
        # for file in files:
        #     #print('------', file)            # File name = file
        #     word = file.split('_')[0][4:]
        #     img  = file.split('_')[1][3:].split('.')[0]
        #     try:
        #         recprob = infer(model, FilePaths.fnInfer + os.path.basename(root) + "/" + file) # Infer the image and get the recognized and probability as a tuple
        #     except:
        #         err += 1
        #         print("Error in line: ", os.path.basename(root), file, "Total errors: ", err)
        #         pass
        #     cellsR [int(word)][int(img)] = recprob[0][0] #store recognized word
        #     cellsP [int(word)][int(img)] = recprob[1][0] #store its probability

            
        # for w in range(len(cellsR)):
        #     wr.append(cellsR[w])
        #     wp.append(cellsP[w])
        # wb.save("inferOutput.xlsx")

    wbb = Workbook()
    for i in range(len(data)):
        wrb = wbb.create_sheet("line"+str(i))
        wpb = wbb.create_sheet("line"+str(i)+"_prob")
        for w in range(len(data[i][0])):
            wrb.append(data[i][0][w])
            wpb.append(data[i][1][w])
    wbb.save("inferOutput2.xlsx")


    wb.save("inferOutput.xlsx")
    print("FINISHED. TOTAL ERRORS: ", err)
            #infer(model, FilePaths.fnInfer)


if __name__ == '__main__':
    main()
